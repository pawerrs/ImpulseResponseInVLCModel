from openpyxl import load_workbook, Workbook
import utils
import os

def save(file_name, values):

    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        page = workbook.active
    else:
        workbook = Workbook()
        page = workbook.active
        page.title = 'Input parameters'
        for i, header in enumerate(utils.get_headers()):
            page.cell(row=1, column=i+1).value = header
    
    max = page.max_row + 1
    for row_index, row in enumerate(values):
        for column_index, column in enumerate(row):
            page.cell(row=max + row_index, column=1 + column_index).value = column
    workbook.save(filename = file_name)
