from openpyxl import load_workbook, Workbook
import utils
import os
from itertools import islice

def handle_data(file_name, values):

    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        training_sheet = workbook.active
        test_sheet = workbook.get_sheet_by_name('Test Dataset')
        validation_sheet = workbook.get_sheet_by_name('Validation Dataset')
    else:
        workbook = Workbook()
        training_sheet = workbook.active
        test_sheet = workbook.create_sheet()
        validation_sheet = workbook.create_sheet()
        training_sheet.title = 'Training Dataset'
        test_sheet.title = 'Test Dataset'
        validation_sheet.title = 'Validation Dataset'
        for i, header in enumerate(utils.get_headers()):
            training_sheet.cell(row=1, column=i+1).value = header
            test_sheet.cell(row=1, column=i+1).value = header
            validation_sheet.cell(row=1, column=i+1).value = header
    
    rows = len(values)
    training_sheet_size = int(0.7*rows)
    validation_sheet_size = int(0.15*rows)
    test_sheet_size = int(0.15*rows)

    save_data(training_sheet, values, 0, training_sheet_size)
    save_data(validation_sheet, values, training_sheet_size, training_sheet_size + validation_sheet_size)
    save_data(test_sheet, values,training_sheet_size + validation_sheet_size + test_sheet_size, None)
    workbook.save(filename = file_name)


def save_data(sheet, values, start_index, end_index):
    max = sheet.max_row + 1
    for row_index, row in enumerate(list(islice(values, start_index, end_index))):
        for column_index, column in enumerate(row):
            sheet.cell(row=max + row_index, column=1 + column_index).value = column


def save(file_name, values):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        page = workbook.active
    else:
        workbook = Workbook()
        page = workbook.active
        page.title = 'Input parameters'
        for i, header in enumerate(utils.get_headers()):
            page.cell(row=1, column=i+1).value = header
    
    max = page.max_row + 1
    for row_index, row in enumerate(values):
        for column_index, column in enumerate(row):
            page.cell(row=max + row_index, column=1 + column_index).value = column
    workbook.save(filename = file_name)
